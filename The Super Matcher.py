# Better Fuzzy Lookup

from thefuzz import fuzz
from thefuzz import process
import jaro
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
import re
import sys
import time
from threading import Thread


#########################################################################################
#                                                                                       #
#                                 The Super Matcher                                     #
#                                                                                       #
#                                By Jack Hinchliffe                                     #
#                             April 2024, Python 3.8.1                                  #
#                                                                                       #
# Requirements:                                                                         #
# - thefuzz 0.20.0                                                                      #
# - jaro-winkler 2.0.3                                                                  #
# - pandas 1.5.0                                                                        #
# - openpyxl 3.0.10                                                                     #
#                                                                                       #
# Features:                                                                             #
# - Launches a GUI for ease of use                                                      #
# - Read a single workbook into the application                                         #
# - Writes left table with matched data from right table if found to loaded workbook    #
# - Significantly faster than the Excel Fuzzy Matching Add-in for large data sets       #
# - Same customizable options as Excel Add-in (complete functional replacement)         #
# - Ability to enable smart matching (self deciding) on the results                     #
# - Self decide suggests which are actual matches based on selected column's similarity #
# - Assists in identifying similar data between large datasets quickly                  #
#                                                                                       #
#########################################################################################


################################################
# Classes
################################################

class TheSuperMatcher(tk.Tk):
    """
    Main window of application gui

    Top level class for script
    """

    # Frame construction
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.title("The Super Matcher")
        self.geometry('1000x750')
        self.frames = {}
        self.minsize(width=750, height=625)

        for F in ({MainPage}):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame(MainPage)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()    

class TextRedirector(object):
    """
    Redirects text from sys output to text widget
    """
    def __init__(self, widget, tag="stdout"):
        self.widget = widget
        self.tag = tag

    def write(self, str):
        self.widget.configure(state="normal")
        self.widget.insert("end", str, (self.tag,))
        self.widget.see("end")
        self.widget.configure(state="disabled")

    def flush(self):
        pass

class Workbook:
    """
    Class for representing an excel workbook within this script

    Parameters
    ----------
    path : str
        a string path to the file that will be represented by this class

    Constructor will create instances of the Tables class for each sheet present in workbook
    """
    matchedText = 'Rows Matched: 0'

    def __init__(self, path:str) -> None:
        self.path = path
        self.tables = {}
        self.__read()
        print('Workbook Loaded!\n')
    
    def __read(self):
        with pd.ExcelFile(self.path) as xls:
            for i, sheet in enumerate(xls.sheet_names):
                print(f'Reading sheet \'{sheet}\'...')
                sheetdata = pd.read_excel(xls, sheet_name=sheet)
                self.tables[sheet] = Tables(sheet, i)
                self.tables[sheet].readData(sheetdata)
#
    def getSheets(self) -> list:
        return list(self.tables.keys())

    def numOfSheets(self) -> int:
        return len(self.tables)

    def getPath(self) -> str:
        return self.path
    
    def setMatchedCount(self, count) -> str:
        self.matchedText = f'Rows Matched: {count}'

    def getMatchedCount(self) -> str:
        return self.matchedText
    

class Tables:
    """
    Tables class that represents a workbook's sheets and the data stored in them

    Parameters
    ----------
    table_name : str
        A string name for this instance table
    table_id : int
        An interger id to index tables
    """
    def __init__(self, table_name, table_id) -> None:
        self.table_name = table_name
        self.table_id = table_id
        self.data = pd.DataFrame()

    def __readColumnHeader(self):
        """
        Gets column names from the sheet and adds sheet name to the column header
        """
        print(f"Extracting columns from \'{self.table_name}\'...\n")
        oldColList = self.data.columns.values.tolist()
        newColList = [f"{col}.{self.table_name}" for col in oldColList]
        oldNewDict = {}
        for old, new in zip(oldColList, newColList):
            oldNewDict[old] = new
        self.data.rename(columns=oldNewDict, inplace=True)
        self.columns = self.data.columns.values.tolist()
        if len(self.columns) == 0: 
            self.columns = ['']

    def readData(self, sheetData: pd.DataFrame) -> None:
        """
        Read the data from the selected sheet
        """
        print("Reading table data...")
        self.data = sheetData
        self.__readColumnHeader()
    
    def getHeaders(self) -> list:
        """
        Get the headers (column name) of this table

        Returns
        -------
        list of column names (str)
        """
        return self.columns
    
    def getData(self) -> pd.DataFrame:
        """
        Get the full data of this table
        """
        return self.data
    
    def getName(self) -> str:
        """
        Get the name of this table
        """
        return self.table_name

################################################
# Helper functions
################################################

def chooseFileHandler(filePathLabel:tk.Label) -> str:
    """
    Prompts user to select a file to load

    Parameters
    ----------
    filePathLabel : tk.Label
        the GUI label that should be updated to display the filepath
    -------
    Returns
    file path string
    """
    excelFile = askopenfilename()
    if(excelFile == ""):
        filePathLabel['text'] = ""
    elif(".xlsx" not in excelFile):
        tk.messagebox.showerror("Error", "Incorrect File Format")
        filePathLabel['text'] = ""
        excelFile = ""
    return excelFile

def writeData(fluData:pd.DataFrame, excelFilePath:str, sheet_name:str, decidedTable:pd.DataFrame, decidedSheet_name:str='') -> None:
    """
    Writes dataframe to an excel sheet of file

    Parameters
    ----------
    data : pandas dataframe
        dataframe to write to file
    excelFilePath : str
        path to excel file
    sheet_name : str
        name of sheet that will be created and wrote to
    """
    print("Updating Workboook (This takes a while!)...\n")
    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False

    name = sheet_name
    dname = decidedSheet_name

    if len(sheet_name) > 30: # Shorten name to prevent breaking workbook. Max is 31 char, but need room incase a number is added for page name duplicates.
        name = sheet_name[:30]
    if len(decidedSheet_name) > 30: # Shorten name to prevent breaking workbook
        dname = decidedSheet_name[:30]

    with pd.ExcelWriter(excelFilePath,
                        mode='a',
                        engine='openpyxl', #using openpyxl because xlsxwritter does not support append mode. Openpyxl is super slow though.
                        if_sheet_exists='new',
    ) as writer:
        writer.workbook = load_workbook(excelFilePath)
        fluData.to_excel(excel_writer=writer, sheet_name=name, index=False)
        if not decidedTable.empty:
            decidedTable.to_excel(excel_writer=writer, sheet_name=dname, index=False)
        print("Data Updated.\nClosing Workbook (This takes a while, too!)...\n") 
    print("Workbook Saved and Closed.\n")

def runButtonHandler(Workbook: Workbook, sheet_selection:list, column_selection:list, similarity_threshold:int, matchLimit:int, doSelfDeciding:bool, matchMode:str, selfDecideParams:dict=None) -> None:
    """
    Handler function to execute main program logic

    Parameters
    -----------
    Workbook : Worbook
        A workbook object instance from this file's workbook class
    sheet_selection : list
        list of strings, length 2
    column_selection : list
        list of strings, length 2
    similarity_threshold : int
        between 0 and 100
    matchLimit : int
        number of matches to find
    doSelfDediciding : bool
        should run self deciding function
    matchMode : str
        matching scorer for fuzzy lookup
    selfDecideParams : dict
        parameter values for the self decide function
        dict {'parameter': [int, str, str, bool]}
    """
    start_time = time.perf_counter()
    tableData = []
    names = []
    decidedTable = pd.DataFrame()

    prefix = 'FLU'

    for sheet in sheet_selection:
        tableData.append(Workbook.tables[sheet].getData())
        names.append(Workbook.tables[sheet].getName())

    if matchMode == "Keyword Search":
        matchedtable = keywordMatch(tableData, column_selection)
        prefix = 'KWS'
    else:
        matchedtable = fuzzyMatch(tableData, column_selection, similarity_threshold, matchLimit)
        prefix = 'FLU'

    if doSelfDeciding and selfDecideParams != None:
        decidedTable = selfDecide(matchedtable, selfDecideParams)

    writeData(matchedtable, Workbook.getPath(), f"{prefix}_{names[0]}_{names[1]}", decidedTable, f'SD_{names[0]}_{names[1]}')
    print(f"Runtime Finished.\nexecution time = {(time.perf_counter() - start_time):.1f} seconds")

#original
def fuzzyMatch(tabledata:list, column_selection:list, similarity_threshold:int, matchLimit:int) -> pd.DataFrame:
    """
    Preforms the fuzzy match between two tables

    Parameters
    ----------
    tabledata : list[pd.DataFrame, pd.DataFrame]
        list of the two dataframes to match
    column_selection : list[str,str]
        list of columns to match on
    similarity_threshold : int
        threshold used to decide whether to reject or keep match result
    matchLimit : int
        maximum number of matches that should be returned
    matchMode : str
        type of fuzzy matching scorer

    Returns
    ----------
    dataframe of left table with matches from right table and a similarity score column
    """
    left_table:pd.DataFrame = tabledata[0].fillna("")
    right_table:pd.DataFrame = tabledata[1].fillna("")
    matchOnLeft = column_selection[0]
    matchOnRight = column_selection[1]

    left_table_len = len(left_table.index)

    scorer = fuzz.token_sort_ratio # Default is token_sort_ratio

    global matchCountText
    matchedCounter = 0
    print('Beginning Fuzzy Matching...\n')

    combinedTable = pd.DataFrame(columns=[*left_table.columns.values, *right_table.columns.values, 'Similarity Score'])
    
    for index, val in left_table.iterrows():
        returnedValues = process.extract(val[matchOnLeft], right_table[matchOnRight], scorer=scorer, limit=matchLimit)
        filteredValues = []
        skippedPairs = 0
        for pair in returnedValues:
            if pair[1] >= similarity_threshold or len(returnedValues) - skippedPairs == 1:
                filteredValues.append(pair)
            else:
                skippedPairs += 1

        for pair in filteredValues:
            if len(filteredValues) == 1 and pair[1] < similarity_threshold:
                pair = ("", 0)
                combinedrow = left_table.iloc[index]
                combinedrow['Similarity Score'] = pair[1]
                combinedTable = combinedTable.append(combinedrow, ignore_index=True) # NOTE using pd.append() is depreciated
            else:
                row_from_left = left_table.iloc[index]
                row_from_right = right_table.iloc[pair[2]]
                combinedrow = pd.concat([row_from_left, row_from_right])
                combinedrow['Similarity Score'] = pair[1]
                combinedTable = combinedTable.append(combinedrow, ignore_index=True) # NOTE using pd.append() is depreciated
                matchedCounter += 1
        matchCountText.set(f'Row Matches Found: {matchedCounter} | Row: {index+1}/{left_table_len}')
    print(f"Fuzzy Matching Complete, {matchedCounter} matches found.\n")
    return combinedTable

def keywordMatch(tabledata:list, colSelection:list) -> pd.DataFrame:
    """
    Match keywords to substrings using regex.
    *Does not accept any limits on matches to return - returns every match.*

    Parameters
    ----------
    tabledata : list[pd.DataFrame, pd.DataFrame]
        list of the two dataframes to match
    column_selection : list[str,str]
        list of columns to match on
    
    Returns
    ----------
    dataframe of left table with matches from right table and a similarity score (100%) column
    """

    left_table:pd.DataFrame = tabledata[0].fillna("")
    right_table:pd.DataFrame = tabledata[1].fillna("")
    matchOnLeft = colSelection[0]
    matchOnRight = colSelection[1]

    left_table_len = len(left_table.index)

    #combinedTable = pd.DataFrame(columns=[*left_table.columns.values, *right_table.columns.values, 'Similarity Score'])
    matchedRows = []

    global matchCountText
    matchedCounter = 0

    print('Beginning Keyword Search...\n')

    for index, val in left_table.iterrows():
        regexString = r'\b{}\b'.format(re.escape(val[matchOnLeft]))
        matches = right_table[right_table[matchOnRight].str.contains(regexString, case=False, regex=True)]

        for _, matchRow in matches.iterrows():
            combinedRow = {**val.to_dict(), **matchRow.to_dict(), 'Similarity Score': 100}
            matchedRows.append(combinedRow)
            matchedCounter += 1
        matchCountText.set(f'Keyword Matches Found: {matchedCounter} | Row: {index+1}/{left_table_len}')
    
    combinedTable = pd.DataFrame(matchedRows)
    print(f"Keyword Matching Complete, {matchedCounter} matches found.\n")
    return combinedTable

def selfDecide(data:pd.DataFrame,  selfDecideParams:dict) -> pd.DataFrame:
    """
    Self deciding logic

    Parameters
    -----------
    data : DataFrame
        Matched table to run decisions on
    selfDecideParams : dict
        parameters to make decisions with
        dict {'parameter': [int, str, str, bool]}
    
    Returns
    -----------
    Inputted Dataframe with calculated decisions in first column
    """

    print("Beginning Self Deciding On Results...\n")

    data = data.fillna('')

    decisionDF = data.copy()
    
    decisionList = []
    for i in range(len(data)):
    
        numMatched = 0
        matchedColNames = []

        for key, value in selfDecideParams.items():
            threshold = value[0]/100
            leftitem = str(data[value[1]].iloc[i]).upper() # To compare str of same case, upper because jaro.original_metric typo checker only compares capitals
            rightitem = str(data[value[2]].iloc[i]).upper()

            if not (leftitem == rightitem == '' and value[3]):
                ratio = jaro.original_metric(leftitem, rightitem)

                if ratio >= threshold and "" not in value:
                    numMatched +=1
                    matchedColNames.append(value[1])
        
        colStr = ', '.join(matchedColNames)

        # TODO: Make this dynamic so that user can select 1-3 criteria instead of requiring all 3
        if numMatched == 0:
            #decisionList.append(rowString + '0/1 match')
            decisionList.append(f'0/3, Not a Match')
        elif numMatched == 1:
            #decisionList.append(rowString + '1/3 match')
            decisionList.append(f"1/3, Possible Match: {colStr}")
        elif numMatched == 2:
            #decisionList.append(rowString + '2/3 match')
            decisionList.append(f"2/3, Likely Match, confirm: {colStr}")
        elif numMatched == 3:
            #decisionList.append(rowString + '3/3 match')
            decisionList.append(f"3/3, Definite Match")
        else:
            decisionList.append(f'Somehow matched more than possible')


    decisionDF.insert(0, "DECISION", decisionList)
    print('Self Deciding Complete.\n')
    return decisionDF

################################################
# GUI Window
################################################

matchCountText = None # NOTE: This is a GLOBAL var that will be used to pass info from calculation threads to the GUI

class MainPage(tk.Frame):

    backgroundcolour = "#b2cfbc"

    def initValues(self):
        """
        Initializes class variables upon initial launch of app
        """
        self.similarity_threshold = 50
        self.num_matches = 1
        self.sheetOptions = []
        self.colOptions_1 = []
        self.colOptions_2 = []
        self.excelFilePath = ""
        self.matchLimit = 1
        self.selected_tables = ["",""]
        self.selected_columns = ["",""]
        self.sheetselectorbox_1_val = ''
        self.sheetselectorbox_2_val = ''
        self.selectedsheet1 = tk.StringVar()
        self.selectedsheet2 = tk.StringVar()
        self.selectedcol1 = tk.StringVar()
        self.selectedcol2 = tk.StringVar()
        self.doSmartMatch = tk.BooleanVar(value=False)
        self.matchCountVar = tk.StringVar()
        self.matchingOptions = ["Match Cell Contents", "Keyword Search"]
        self.matchingMode = tk.StringVar(value=self.matchingOptions[0])
        global matchCountText
        matchCountText = tk.StringVar()
        matchCountText.set('Waiting to run...')
        

    def matchingThreadFnc(self):
        """
        This is executed in seperate running/calculation thread, seperate from GUI thread.

        This thread is for running the background calculations for fuzzy match and self decide without freezing the GUI
        """
        self.button_run['state'] = 'disabled'
        runButtonHandler(self.thisWorkbook, self.selected_tables, self.selected_columns, self.similarity_threshold, self.matchLimit, self.doSmartMatch.get(), self.matchingMode.get(), self.compressSelfDecideParams(self.sd_slider, self.sd_combobox, self.sd_compareEmpty))
        self.button_run['state'] = 'normal'

    def onRunPress(self):
        """
        This is called and executed by clicking the 'Run Program' button

        Creates a seperate thread for background calculations!
        - Specifically the matchingThreadFnc thread
        """
        self.similarity_threshold = int(self.slider_similarity.get())
        self.matchLimit = int(self.spinbox_limit.get())
        self.selected_columns[0] = self.combobox_col_selector_1.get()
        self.selected_columns[1] = self.combo_col_selector_2.get()
        self.selected_tables[0] = self.combobox_sheetselector_1.get()
        self.selected_tables[1] = self.combobox_sheetselector_2.get()

        matchingThread = Thread(name='MatchingThread', daemon=True, target=self.matchingThreadFnc)
        matchingThread.start()

    def compressSelfDecideParams(self, slider, listbox, noMatchBlank):
        """
        Squishes the three seperate dicts into a single dict of paramter and values

        Returns either single dict of the parameters or None if not running self decide function
        """
        if self.doSmartMatch.get():
            dataForDecider = {
                "Parameter 1": [int, str, str, bool],
                "Parameter 2": [int, str, str, bool],
                "Parameter 3": [int, str, str, bool]
            }
            for key in dataForDecider:
                dataForDecider[key] = [slider[key][0].get(), listbox[key][0].get(), listbox[key][2].get(), noMatchBlank[key][0].get()]
        else:
            dataForDecider = None
        
        return dataForDecider

    def chooseThreadFnc(self):
        """
        This is executed in seperate running/calculation thread, seperate from GUI thread.

        This thread is for reading the worbook and creating this script's objects without freezing the GUI.
        """
        self.thisWorkbook = Workbook(self.excelFilePath)
        self.sheetOptions = self.thisWorkbook.getSheets()
        

        self.combobox_sheetselector_1['values'] = self.sheetOptions
        self.combobox_sheetselector_2['values'] = self.sheetOptions

        self.selectedsheet1 = self.sheetOptions[0]

        if self.thisWorkbook.numOfSheets() < 2:
            self.selectedsheet2 = self.selectedsheet1
        else:
            self.selectedsheet2 = self.sheetOptions[1]

        self.sheetselectorbox_1_val = self.selectedsheet1
        self.sheetselectorbox_2_val = self.selectedsheet2
        self.combobox_sheetselector_1.set(self.selectedsheet1)
        self.combobox_sheetselector_2.set(self.selectedsheet2)
        
        self.combobox_col_selector_1.set(self.thisWorkbook.tables[self.selectedsheet1].getHeaders()[0])
        self.combo_col_selector_2.set(self.thisWorkbook.tables[self.selectedsheet2].getHeaders()[0])
        self.combobox_col_selector_1['values'] = self.thisWorkbook.tables[self.selectedsheet1].getHeaders()
        self.combo_col_selector_2['values'] = self.thisWorkbook.tables[self.selectedsheet2].getHeaders()

        # TODO: Remove these repeat bind assignments... They're assigned in __init__, why did I do this here in the first place??
        self.combobox_sheetselector_1.bind('<<ComboboxSelected>>', self.onSheetSelect_1)
        self.combobox_sheetselector_2.bind('<<ComboboxSelected>>', self.onSheetSelect_2)
        self.combobox_col_selector_1.bind('<<ComboboxSelected>>', self.checkToEnableRun)
        self.combo_col_selector_2.bind('<<ComboboxSelected>>', self.checkToEnableRun)

        self.popSelfDecideWidgets()
        self.checkToEnableRun()
        self.button_chooseFile['state'] = 'normal'
        self.info["state"] = "disabled"
        global matchCountText
        matchCountText.set('File Loaded, Waiting to run')

    def onChooseFilePress(self):
        """
        This is called and executed by clicking the 'Choose File' button

        Creates a seperate thread for background calculations!
        - Specifically the chooseThreadFnc thread
        """
        global matchCountText
        self.button_chooseFile['state'] = 'disabled'
        self.excelFilePath = chooseFileHandler(tk.Label(self))
        if self.excelFilePath != '':

            self.text_sourcefile_disp['state'] = 'normal'
            self.text_sourcefile_disp.delete('1.0', 'end')
            self.text_sourcefile_disp.insert("end", self.excelFilePath)
            self.text_sourcefile_disp['state'] = 'disabled'

            matchCountText.set('Loading Workbook, please wait...')
            print(f'Reading Workbook \'{self.excelFilePath}\'')

            creatingWorkbookThread = Thread(target=self.chooseThreadFnc, name='ChooseThread', daemon=True)
            creatingWorkbookThread.start()
        else:
            self.button_chooseFile['state'] = 'normal'
        

    def onSheetSelect_1(self, event):
        """
        Event handler for what to do when the Left Table is selected from drop down
        """
        if self.excelFilePath == '': return
        self.sheetselectorbox_1_val = event.widget.get()
        #self.colOptions_1 = readColumns(self.excelFile, self.sheetselectorbox_1_val)
        self.colOptions_1 = self.thisWorkbook.tables[self.sheetselectorbox_1_val].getHeaders()
        self.combobox_col_selector_1['values'] = self.colOptions_1
        self.selectedcol1 = self.colOptions_1[0]
        self.combobox_col_selector_1.set(self.selectedcol1)
        self.checkToEnableRun()
        self.popSelfDecideWidgets()

    def onSheetSelect_2(self, event):
        """
        Event handler for what to do when the Right Table is selected from drop down
        """
        if self.excelFilePath == '': return
        self.sheetselectorbox_2_val = event.widget.get()
        #self.colOptions_2 = readColumns(self.excelFile, self.sheetselectorbox_2_val)
        self.colOptions_2 = self.thisWorkbook.tables[self.sheetselectorbox_2_val].getHeaders()
        self.combo_col_selector_2['values'] = self.colOptions_2
        self.selectedcol2 = self.colOptions_2[0]
        self.combo_col_selector_2.set(self.selectedcol2)
        self.checkToEnableRun()
        self.popSelfDecideWidgets()

    def checkToEnableRun(self, event=None):
        """
        Changes run button state by checking if all required info has been entered to enable program to run
        """
        allSelected = False

        for key in self.sd_combobox:
            for i in range(len(self.sd_combobox[key])):
                if type(self.sd_combobox[key][i]) == ttk.Combobox:
                    if self.sd_combobox[key][i].get():
                        allSelected = True
                    else:
                        allSelected = False

        if self.excelFilePath != '':
            if (self.combobox_sheetselector_1.get() != self.combobox_sheetselector_2.get()) and (self.combobox_col_selector_1.get() != '' or self.combo_col_selector_2.get() != ''):
                if self.doSmartMatch.get():
                    if allSelected:
                        self.button_run['state'] = 'normal'
                    else:
                        self.button_run['state'] = 'disabled'
                else:
                    self.button_run['state'] = 'normal'
            else:
                self.button_run['state'] = 'disabled'

    def changeWidgetCollectionState(self, obj:dict, state:str):
        """
        Changes state (enabled/disabled) of the widgets in dict based off parameter passed
        """
        for key in obj:
            for i in range(len(obj[key])):
                if type(obj[key][i]) == ttk.Combobox:
                        obj[key][i]['state'] = state
                        obj[key][i].state(['readonly'])
                elif isinstance(obj[key][i], tk.Widget):
                        obj[key][i]['state'] = state
    
    def unlockSelfDecideWidgets(self, event=None):
        """
        Event handler that changes state of self-decide related widgets if checkbox is clicked on/off
        """
        if self.doSmartMatch.get():
            self.changeWidgetCollectionState(self.sd_combobox, 'normal')
            self.changeWidgetCollectionState(self.sd_slider, 'normal')
            self.changeWidgetCollectionState(self.sd_compareEmpty, 'normal')
            self.popSelfDecideWidgets()
        else:
            self.changeWidgetCollectionState(self.sd_combobox, 'disabled')
            self.changeWidgetCollectionState(self.sd_slider, 'disabled')
            self.changeWidgetCollectionState(self.sd_compareEmpty, 'disabled')
    
    def toggleFuzzyWidgets(self, event:tk.Event):
        mode = event.widget.get()
        if mode == 'Keyword Search':
            self.slider_similarity['state'] = 'disabled'
            self.spinbox_limit['state'] = 'disabled'
            self.label_similarity['state'] = 'disabled'
            self.label_slider['state'] = 'disabled'
        else:
            self.slider_similarity['state'] = 'normal'
            self.spinbox_limit['state'] = 'normal'
            self.label_similarity['state'] = 'normal'
            self.label_slider['state'] = 'normal'

    def clearlistbox(self, listbox):
        """
        Clears the listbox of values
        
        Depreciated use since converting listboxes to combobox
        """
        for key in listbox:
            for i in range(len(listbox[key])):
                if type(listbox[key][i]) == tk.Listbox:
                    listbox[key][i].delete(0, listbox[key][i].size())

    def popSelfDecideWidgets(self):
        """
        Populates the self decide widgets based on left/right table selections.
        """
        self.checkToEnableRun()
        vals1 = self.combobox_col_selector_1['values']
        vals2 = self.combo_col_selector_2['values']
        try: # Try to populate the listbox. If workbook object hasn't been created yet (no file loaded), pass on AttributeError
            if self.doSmartMatch.get():
                for key in self.sd_combobox:
                    self.sd_combobox[key][0]['values'] = vals1
                    self.sd_combobox[key][0].set(vals1[0])
                    self.sd_combobox[key][2]['values'] = vals2
                    self.sd_combobox[key][2].set(vals2[0])
        except AttributeError:
            pass

    def __init__(self, parent, controller):
        """
        MainPage Frame constructor overloading.

        This is where widgets and tkinter vars are created and initialized.
        """

        global matchCountText
        tk.Frame.__init__(self, parent)
        self.initValues()


        ################################
        # Row 0
        ################################
        header = tk.Label(self, text="The Super Matcher", font="Arial 20 bold")
        header.configure(background=self.backgroundcolour)
        header.grid(column=1, row=0, pady=10)

        ################################
        # Row 1
        ################################

        self.button_chooseFile = tk.Button(self, text="Choose File", font="Arial 14", command=self.onChooseFilePress, cursor='hand2')
        self.button_chooseFile.grid(row=2, column=0)

        self.button_run = tk.Button(self, text="Start Matching", font="Arial 14", command=self.onRunPress, cursor='hand2')
        self.button_run.grid(row=2, column=2)
        self.button_run['state'] = 'disabled'

        ################################
        # Row 2
        ################################
        
        self.label_sourcefile = tk.Label(self)
        self.label_sourcefile.configure(background=self.backgroundcolour, font='Roboto 10 bold')
        self.label_sourcefile.grid(column=0, row=3,sticky="ne", padx=5, pady=5)
        self.label_sourcefile['text'] = 'File Path:'

        self.text_sourcefile_disp = tk.Text(self)
        self.text_sourcefile_disp.configure(height=1, font='arial 10')
        self.text_sourcefile_disp.grid(column=1, row=3,sticky="nwe",pady=5, columnspan=2, padx=15)
        self.text_sourcefile_disp['state'] = 'disabled'

        ################################
        # Row 3
        ################################

        self.slider_similarity = tk.Scale(self, from_=0, to=100, orient="horizontal", cursor='hand2')
        self.slider_similarity.grid(row = 6, column=1)
        self.slider_similarity.set(self.similarity_threshold)

        self.spinbox_limit = tk.Spinbox(self, from_=1, to=10) #Max 10 matches allowed
        self.spinbox_limit.grid(row=8, column=1)

        ################################
        # Row 4
        ################################
        
        #combo boxes
        self.combobox_sheetselector_1 = ttk.Combobox(self, textvariable=self.selectedsheet1, width=30)
        self.combobox_sheetselector_1['values'] = self.sheetOptions
        self.combobox_sheetselector_1.state(["readonly"])
        self.combobox_sheetselector_1.grid(row=6, column=0)

        self.combobox_sheetselector_2 = ttk.Combobox(self, textvariable=self.selectedsheet2, width=30)
        self.combobox_sheetselector_2['values'] = self.sheetOptions
        self.combobox_sheetselector_2.state(["readonly"])
        self.combobox_sheetselector_2.grid(row=6, column=2)

        self.label_L_table = tk.Label(self)
        self.label_L_table.configure(background=self.backgroundcolour)
        self.label_L_table.grid(column=0, row=5,sticky="nw", padx=15)
        self.label_L_table['text'] = 'Left Table (find matches for)'

        self.label_R_table = tk.Label(self)
        self.label_R_table.configure(background=self.backgroundcolour)
        self.label_R_table.grid(column=2, row=5,sticky="nw", padx=15)
        self.label_R_table['text'] = 'Right Table (find matches from)'

        self.label_L_column = tk.Label(self)
        self.label_L_column.configure(background=self.backgroundcolour)
        self.label_L_column.grid(column=0, row=7,sticky="nw", padx=15)
        self.label_L_column['text'] = 'Left Matching Column'

        self.label_R_column = tk.Label(self)
        self.label_R_column.configure(background=self.backgroundcolour)
        self.label_R_column.grid(column=2, row=7,sticky="nw", padx=15)
        self.label_R_column['text'] = 'Right Matching Column'

        self.combobox_col_selector_1 = ttk.Combobox(self, textvariable=self.selectedcol1, width=30)
        self.combobox_col_selector_1['values'] = self.colOptions_1
        self.combobox_col_selector_1.state(["readonly"])
        self.combobox_col_selector_1.bind('<<ComboboxSelected>>', self.checkToEnableRun)
        self.combobox_col_selector_1.grid(row=8, column=0, padx=15)

        self.combo_col_selector_2 = ttk.Combobox(self, textvariable=self.selectedcol2, width=30)
        self.combo_col_selector_2['values'] = self.colOptions_2
        self.combo_col_selector_2.state(["readonly"])
        self.combo_col_selector_2.bind('<<ComboboxSelected>>', self.checkToEnableRun)
        self.combo_col_selector_2.grid(row=8, column=2, padx=15)

        self.label_similarity = tk.Label(self)
        self.label_similarity.configure(background=self.backgroundcolour)
        self.label_similarity.grid(column=1, row=5,sticky="nwe")
        self.label_similarity['text'] = 'Similarity Cutoff'

        self.label_slider = tk.Label(self)
        self.label_slider.configure(background=self.backgroundcolour)
        self.label_slider.grid(column=1, row=7,sticky="nwe")
        self.label_slider['text'] = 'Maximum Matches to return'

        self.combobox_matchTypeSelector = ttk.Combobox(self, textvariable=self.matchingMode, width=20)
        self.combobox_matchTypeSelector['values'] = self.matchingOptions
        self.combobox_matchTypeSelector.state(['readonly'])
        self.combobox_matchTypeSelector.bind('<<ComboboxSelected>>', self.toggleFuzzyWidgets)
        self.combobox_matchTypeSelector.grid(row=10, column=1, pady=5)

        #self.matchCountVar.set(f'Rows Matched: {matchCount}')
        self.label_matchedCount = tk.Label(self, textvariable=matchCountText)
        self.label_matchedCount.configure(background=self.backgroundcolour, font='TkDefaultFont 10 bold')
        self.label_matchedCount.grid(column=0, row=10,sticky="nw", padx=15)
        

        header_fuzzy = tk.Label(self, text="Fuzzy Matching Options", font="Arial 12 bold")
        header_fuzzy.configure(background=self.backgroundcolour)
        header_fuzzy.grid(column=1, row=1, pady=10)

        # Decider widgets
        header_decide = tk.Label(self, text="Self Deciding Match Options", font="Arial 12 bold")
        header_decide.configure(background=self.backgroundcolour)
        header_decide.grid(column=1, row=11, pady=10)

        self.checkbutton_runSDFuncs = tk.Checkbutton(self, text='Apply Smart Matching Decisions To Results', font="Arial 10 bold", variable=self.doSmartMatch, onvalue=True, offvalue=False, command=self.unlockSelfDecideWidgets, cursor='hand2')
        self.checkbutton_runSDFuncs.configure(background=self.backgroundcolour)
        self.checkbutton_runSDFuncs.grid(column=1, row=12, pady=5)

        self.sd_slider = {"Parameter 1":[], "Parameter 2":[], "Parameter 3":[]}
        self.sd_combobox = {"Parameter 1":[], "Parameter 2":[], "Parameter 3":[]}
        self.sd_compareEmpty = {"Parameter 1":[], "Parameter 2":[], "Parameter 3":[]}

        c=0
        for key in self.sd_slider:
            self.sd_slider[key].append(tk.Scale(self, orient=tk.HORIZONTAL, cursor='hand2'))
            self.sd_slider[key][0].grid(column=c, row=14)
            self.sd_slider[key][0].set(75)
            self.sd_slider[key].append(tk.Label(self, text=key))
            self.sd_slider[key][1].configure(bg=self.backgroundcolour)
            self.sd_slider[key][1].grid(column=c, row=13)
            c += 1
        c=0

        for key in self.sd_combobox:
            self.sd_combobox[key].append(ttk.Combobox(self, width=30))
            self.sd_combobox[key][0].grid(column=c, row=16, pady=5)
            self.sd_combobox[key][0].bind('<<ComboboxSelected>>', self.checkToEnableRun)
            self.sd_combobox[key][0].state(["readonly"])
            self.sd_combobox[key].append(tk.Label(self, text=key, font=("Boulder", 13)))
            self.sd_combobox[key].append(ttk.Combobox(self, width=30))
            self.sd_combobox[key][2].grid(column=c, row=21, pady=5)
            self.sd_combobox[key][2].bind('<<ComboboxSelected>>', self.checkToEnableRun)
            self.sd_combobox[key][2].state(["readonly"])
            c += 1
        c=0
        for key in self.sd_compareEmpty:
            self.sd_compareEmpty[key].append(tk.BooleanVar(value=False))
            self.sd_compareEmpty[key].append(tk.Checkbutton(self, text="Ignore matching blank values", variable=self.sd_compareEmpty[key][0], onvalue=True, offvalue=False, cursor='hand2'))
            self.sd_compareEmpty[key][1].grid(column=c, row=15)
            self.sd_compareEmpty[key][1].configure(background=self.backgroundcolour)
            c += 1

        self.changeWidgetCollectionState(self.sd_combobox, 'disabled')
        self.changeWidgetCollectionState(self.sd_slider, 'disabled')
        self.changeWidgetCollectionState(self.sd_compareEmpty, 'disabled')

        self.info = tk.Text(self, height=5)
        self.info.configure(bg=self.backgroundcolour)
        self.info.grid(column=0, row=26, pady=10, padx=15, columnspan=3, sticky='ew')
        sys.stdout = TextRedirector(self.info, "stdout")
        self.info.see('end')
        

        self.configure(background=self.backgroundcolour)

        self.grid_columnconfigure(0, pad=10)
        self.grid_columnconfigure(1, weight=3, minsize=280)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(11, weight=1)


# Run the app mainloop
if __name__ == '__main__':
    app = TheSuperMatcher()
    app.mainloop()    